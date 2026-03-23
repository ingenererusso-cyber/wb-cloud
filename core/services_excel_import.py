from __future__ import annotations

from datetime import datetime
from typing import Any

from django.utils import timezone

from core.models import Order, SellerAccount
from core.services.localization import determine_locality, normalize_district

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - dependency error is handled in runtime
    load_workbook = None


HEADER_ALIASES = {
    "supplier_article": {"артикул продавца"},
    "nm_id": {"артикул wb", "артикул вб"},
    "order_date": {"дата оформления заказа"},
    "last_change_date": {"дата текущего статуса"},
    "status": {"статус заказа"},
    "finished_price": {"стоимость"},
    "srid": {"id заказа"},
    "warehouse_type": {"тип склада"},
    "region_departure": {"регион отправки"},
    "region_arrival": {"регион прибытия"},
    "tech_size": {"размер wb", "размер"},
}

# Стандартные колонки WB-листа "Все заказы" (A=0, B=1, ...),
# если строка заголовков повреждена/не найдена.
FALLBACK_WB_ALL_ORDERS_MAP = {
    "supplier_article": 0,      # A
    "nm_id": 1,                 # B
    "order_date": 7,            # H
    "last_change_date": 8,      # I
    "status": 9,                # J
    "region_departure": 12,     # M
    "region_arrival": 14,       # O
    "finished_price": 16,       # Q
    "tech_size": 18,            # S
    "srid": 20,                 # U
    "warehouse_type": 21,       # V
}


def _normalize_header(value: Any) -> str:
    text = (str(value or "")).strip().lower().replace("ё", "е")
    return " ".join(text.split())


def _build_header_map(row_values: list[Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, raw in enumerate(row_values):
        normalized = _normalize_header(raw)
        if not normalized:
            continue
        for field, aliases in HEADER_ALIASES.items():
            if normalized in aliases and field not in mapping:
                mapping[field] = idx
    return mapping


def _is_header_map_usable(mapping: dict[str, int]) -> bool:
    # supplier_article может отсутствовать, восстановим его из nm_id.
    return {"nm_id", "order_date"}.issubset(mapping)


def _find_header_row_and_map(rows: list[tuple[Any, ...]]) -> tuple[int | None, dict[str, int]]:
    for idx, row in enumerate(rows[:100]):
        candidate = _build_header_map(list(row or []))
        if _is_header_map_usable(candidate):
            return idx, candidate
    return None, {}


def _parse_dt(raw: Any):
    if not raw:
        return None
    if isinstance(raw, datetime):
        dt = raw
    else:
        text = str(raw).strip()
        if not text:
            return None
        dt = None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%d.%m.%Y %H:%M:%S", "%Y-%m-%d", "%d.%m.%Y"):
            try:
                dt = datetime.strptime(text, fmt)
                break
            except ValueError:
                continue
        if dt is None:
            return None
    if timezone.is_naive(dt):
        return timezone.make_aware(dt, timezone.get_default_timezone())
    return dt


def _parse_float(raw: Any) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip().replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _parse_nm_id(raw: Any) -> int | None:
    if raw is None:
        return None
    if isinstance(raw, int):
        return raw
    if isinstance(raw, float):
        return int(raw)
    text = str(raw).strip()
    if not text:
        return None
    if text.endswith(".0"):
        text = text[:-2]
    try:
        return int(text)
    except ValueError:
        return None


def _get_cell(row_values: list[Any], idx: int | None) -> Any:
    if idx is None:
        return None
    if idx < 0 or idx >= len(row_values):
        return None
    return row_values[idx]


def _is_cancel_status(status_text: str) -> bool:
    normalized = (status_text or "").strip().lower().replace("ё", "е")
    return ("отказ" in normalized) or ("отмен" in normalized)


def _normalize_warehouse_type(value: Any) -> str:
    text = (str(value or "")).strip().lower().replace("ё", "е")
    if "wb" in text:
        return "Склад WB"
    if "маркетплейс" in text or "fbs" in text:
        return "Маркетплейс"
    return str(value or "").strip() or "Склад WB"


def _make_synthetic_srid(row_index: int, nm_id: int, order_date) -> str:
    date_part = order_date.strftime("%Y%m%d%H%M%S") if order_date else "no_date"
    return f"excel-{nm_id}-{date_part}-{row_index}"


def _determine_locality_from_excel_regions(
    region_departure: str | None,
    region_arrival: str | None,
) -> bool | None:
    departure = normalize_district(region_departure)
    arrival = normalize_district(region_arrival)
    if not departure or not arrival:
        return None
    return departure == arrival


def import_orders_from_excel(
    *,
    seller: SellerAccount,
    file_obj,
) -> dict[str, int]:
    """
    Импортирует заказы из XLSX-выгрузки WB (лист "Все заказы") в Order.

    Данные неполные: используются только поля, доступные в файле.
    При последующей API-синхронизации записи будут дозаполнены/обновлены.
    """
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed")

    # Для некоторых выгрузок WB размер листа в metadata поврежден.
    # В режиме read_only openpyxl тогда отдает только первую строку.
    # Поэтому читаем в обычном режиме.
    workbook = load_workbook(filename=file_obj, read_only=False, data_only=True)

    # Ищем данные по всем листам, начиная с "Все заказы" (в твоём файле это 3-й лист).
    candidate_sheet_names: list[str] = []
    if "Все заказы" in workbook.sheetnames:
        candidate_sheet_names.append("Все заказы")
    candidate_sheet_names.extend([name for name in workbook.sheetnames if name not in candidate_sheet_names])

    rows: list[tuple[Any, ...]] = []
    header_row_idx: int | None = None
    header_map: dict[str, int] = {}

    for sheet_name in candidate_sheet_names:
        sheet_rows = list(workbook[sheet_name].iter_rows(values_only=True))
        if not sheet_rows:
            continue
        found_idx, found_map = _find_header_row_and_map(sheet_rows)
        if found_idx is not None:
            rows = sheet_rows
            header_row_idx = found_idx
            header_map = found_map
            break

    # Fallback на фиксированную карту WB-колонок.
    if header_row_idx is None:
        if "Все заказы" in workbook.sheetnames:
            rows = list(workbook["Все заказы"].iter_rows(values_only=True))
        elif workbook.sheetnames:
            rows = list(workbook[workbook.sheetnames[0]].iter_rows(values_only=True))
        if rows:
            header_row_idx = 1 if len(rows) > 1 else 0
            header_map = FALLBACK_WB_ALL_ORDERS_MAP.copy()

    if header_row_idx is None or not rows:
        raise ValueError("Не удалось найти данные заказов в Excel")

    # В выгрузке "Все заказы" склад отправки идёт следующим столбцом после "Регион отправки".
    warehouse_name_idx = None
    if "region_departure" in header_map:
        warehouse_name_idx = header_map["region_departure"] + 1

    created = 0
    updated = 0
    skipped = 0

    for rel_row_idx, row in enumerate(rows[header_row_idx + 1 :], start=header_row_idx + 2):
        values = list(row or [])
        if not any(values):
            continue

        nm_id = _parse_nm_id(_get_cell(values, header_map.get("nm_id")))
        supplier_article = str(_get_cell(values, header_map.get("supplier_article")) or "").strip()
        if not nm_id:
            skipped += 1
            continue
        if not supplier_article:
            supplier_article = f"nm_{nm_id}"

        order_date = _parse_dt(_get_cell(values, header_map.get("order_date")))
        last_change_date = _parse_dt(_get_cell(values, header_map.get("last_change_date"))) or order_date
        if order_date is None:
            skipped += 1
            continue
        if last_change_date is None:
            last_change_date = timezone.now()

        srid = str(_get_cell(values, header_map.get("srid")) or "").strip()
        if not srid:
            srid = _make_synthetic_srid(rel_row_idx, nm_id, order_date)

        status_text = str(_get_cell(values, header_map.get("status")) or "").strip()
        is_cancel = _is_cancel_status(status_text)
        warehouse_name = str(_get_cell(values, warehouse_name_idx) or "").strip()
        warehouse_type = _normalize_warehouse_type(_get_cell(values, header_map.get("warehouse_type")))
        oblast_okrug_name = str(_get_cell(values, header_map.get("region_arrival")) or "").strip() or None
        region_name = str(_get_cell(values, header_map.get("region_departure")) or "").strip() or None
        is_local: bool
        if warehouse_type == "Склад WB":
            # Для Excel-импорта сначала используем прямое сравнение "регион отгрузки vs регион доставки".
            # Если регионы отсутствуют, откатываемся на сопоставление по складу.
            region_based_locality = _determine_locality_from_excel_regions(region_name, oblast_okrug_name)
            if region_based_locality is None:
                is_local = determine_locality(warehouse_name, oblast_okrug_name or "")
            else:
                is_local = region_based_locality
        else:
            is_local = False

        finished_price = _parse_float(_get_cell(values, header_map.get("finished_price")))
        tech_size = str(
            _get_cell(values, header_map.get("tech_size")) or ""
        ).strip()[:100]

        _, was_created = Order.objects.update_or_create(
            seller=seller,
            srid=srid,
            defaults={
                "nm_id": nm_id,
                "supplier_article": supplier_article[:255],
                "tech_size": tech_size,
                "warehouse_name": warehouse_name[:255],
                "warehouse_type": warehouse_type[:50],
                "country_name": "Россия",
                "oblast_okrug_name": (oblast_okrug_name or "")[:255] or None,
                "region_name": (region_name or "")[:255] or None,
                "is_cancel": is_cancel,
                "finished_price": finished_price,
                "order_date": order_date,
                "last_change_date": last_change_date,
                "is_local": is_local,
            },
        )
        if was_created:
            created += 1
        else:
            updated += 1

    return {"created": created, "updated": updated, "skipped": skipped}
